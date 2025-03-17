import xlsxwriter as xr
from openpyxl.reader.excel import load_workbook
from DoAnCuoiKy.models.Material import Material


class Tool:
    def export_materials_to_excel(self,filename,materials):
    #export danh sách nguyên liệu
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 15)
        worksheet.set_column('F:F', 15)
        worksheet.set_column('G:G', 15)
        worksheet.set_column('H:H', 15)
        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'Name', bold)
        worksheet.write('B1', 'Type', bold)
        worksheet.write('C1', 'Supplier', bold)
        worksheet.write('D1', 'Import Price', bold)
        worksheet.write('E1', 'Import Quantity', bold)
        worksheet.write('F1', 'Sell Price', bold)
        worksheet.write('G1', 'Sold Quantity', bold)
        worksheet.write('H1', 'Use Day', bold)
        for i in range(len(materials)):
            index = i + 2
            mt = materials[i]
            worksheet.write(f'A{index}', mt.name)
            worksheet.write(f'B{index}', mt.type)
            worksheet.write(f'C{index}', mt.supplier)
            worksheet.write(f'D{index}', mt.import_price)
            worksheet.write(f'E{index}', mt.import_qty)
            worksheet.write(f'F{index}', mt.sell_price)
            worksheet.write(f'G{index}', mt.sold_qty)
            worksheet.write(f'H{index}', mt.use_day)
        workbook.close()

    def export_employees_to_excel(self,filename,employees):
    #export danh sách nhân viên
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'Employee ID', bold)
        worksheet.write('B1', 'Employee Name', bold)
        worksheet.write('C1', 'Username', bold)
        worksheet.write('D1', 'Password', bold)
        for i in range(len(employees)):
            index = i + 2
            e = employees[i]
            worksheet.write(f'A{index}', e.EmployeeId)
            worksheet.write(f'B{index}', e.EmployeeName)
            worksheet.write(f'C{index}', e.UserName)
            worksheet.write(f'D{index}', e.Password)
        workbook.close()

    def export_managers_to_excel(self,filename,managers):
    #export danh sách quản lí
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        bold = workbook.add_format({'bold': True})
        worksheet.write('A1', 'Manager ID', bold)
        worksheet.write('B1', 'Manager Name', bold)
        worksheet.write('C1', 'Username', bold)
        worksheet.write('D1', 'Password', bold)
        for i in range(len(managers)):
            index = i+2
            m = managers[i]
            worksheet.write(f'A{index}', m.EmployeeId)
            worksheet.write(f'B{index}', m.EmployeeName)
            worksheet.write(f'C{index}', m.UserName)
            worksheet.write(f'D{index}', m.Password)
        workbook.close()

    def import_material_from_excel(self, filename):
    #nhập dữ liệu từ file excel
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        is_header = True
        materials = []
        for row in ws.values:
            if is_header == True:
                is_header = False
                continue
            name = row[0]
            type = row[1]
            supplier = row[2]
            import_price = row[3]
            import_qty = row[4]
            sell_price = row[5]
            sold_qty = row[6]
            use_day = row[7]
            m = Material(name, type, supplier, import_price, import_qty, sell_price, sold_qty, use_day)
            materials.append(m)
        wb.close()
        return materials
